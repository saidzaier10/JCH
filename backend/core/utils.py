import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from datetime import datetime

def export_registrations_to_excel(registrations):
    """
    Génère un fichier Excel contenant la liste détaillée des inscriptions.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inscriptions"

    # Headers
    headers = [
        "Nom", "Prénom", "Date de Naissance", "Genre", "Adresse", 
        "Catégorie", "Poids", "Discipline", "Ceinture", "Licence", 
        "Droit Image", "Payé", "Statut", "Mode Paiement", "Mensualités", 
        "Aide Mairie", "Montant Aide", "Discipline Suppl.", "Remise (€)",
        "Nom Parent", "Email Parent", "Téléphone", "Saison"
    ]
    
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f"{col_letter}1"] = header
        ws[f"{col_letter}1"].font = openpyxl.styles.Font(bold=True)

    # Data
    for row_num, reg in enumerate(registrations, 2):
        member = reg.member
        parent = member.parent
        
        # Parent Info
        parent_name = f"{parent.first_name} {parent.last_name}" if parent else "N/A"
        parent_email = parent.email if parent else member.email
        
        # Format Boolean
        def format_bool(val):
            return "Oui" if val else "Non"

        # Mapping data to columns
        data = [
            member.last_name,
            member.first_name,
            member.birth_date.strftime("%d/%m/%Y") if member.birth_date else "",
            member.get_gender_display(),
            member.address,
            
            reg.category.name,
            member.get_weight_category_display() if member.weight_category else "",
            member.get_discipline_display(),
            member.get_belt_display(),
            member.license_number,
            
            format_bool(member.image_rights),
            format_bool(reg.paid),
            reg.get_status_display(),
            reg.get_payment_mode_display(),
            reg.installments_paid,
            
            format_bool(reg.city_hall_aid),
            reg.city_hall_aid_amount,
            format_bool(reg.has_supplementary_discipline),
            reg.discount_amount,
            
            parent_name,
            parent_email,
            member.phone,
            reg.season.name
        ]

        for col_num, cell_value in enumerate(data, 1):
            col_letter = get_column_letter(col_num)
            ws[f"{col_letter}{row_num}"] = cell_value

    # Auto-adjust column width
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename=inscriptions_export_{datetime.now().strftime("%Y%m%d")}.xlsx'
    
    wb.save(response)
    return response
